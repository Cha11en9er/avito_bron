from __future__ import annotations

import base64
import io
import json
import os
import random
import re
import sys
import traceback
from datetime import datetime
from pathlib import Path
from parser.paths import project_root
from urllib.parse import urljoin, urlparse

import ddddocr
from PIL import Image, ImageOps
from openpyxl import Workbook, load_workbook
from playwright.sync_api import sync_playwright
from playwright_stealth.stealth import Stealth

NAV_TIMEOUT_MS = 90000
# Как в avito-parser-exactly.py: после domcontentloaded даём SPA/стилям время (сек.).
POST_GOTO_WAIT_S = 30
BETWEEN_SITES_PAUSE_S = 30
URLS_FILE_NAME = "urls.txt"
OUTPUT_DIR_NAME = "avito_houses_dump"
RESULTS_XLSX_FILE_NAME = "дома авито.xlsx"
NOT_FOUND_VALUE = "нету на сайте"
PHONE_NOT_RECOGNIZED = "не распознан"

_OCR = ddddocr.DdddOcr(show_ad=False)

EXPORT_COLUMNS = [
    "номер",
    "название",
    "цена",
    "адрес",
    "телефон",
    "площадь дома",
    "площадь участка",
    "этажей",
    "материал стен",
    "год постройки",
    "описание",
    "характеристики",
    "ссылка",
]

HOUSE_FIELD_KEYWORDS: dict[str, list[str]] = {
    "площадь дома":    ["площадь дома", "общая площадь"],
    "площадь участка": ["площадь участка"],
    "этажей":          ["этажей в доме", "этажност", "этажей"],
    "материал стен":   ["материал стен", "тип дома"],
    "год постройки":   ["год постройки"],
}

# Как в 24.04.26-avito / avito-parser-exactly.py — только разворот окна, без лишних флагов.
CHROME_ARGS = [
    "--start-maximized",
]


def _setup_playwright_env() -> None:
    if getattr(sys, "frozen", False):
        base = Path(sys.executable).resolve().parent
    else:
        base = project_root()
    bundled = base / "ms-playwright"
    if bundled.is_dir():
        os.environ["PLAYWRIGHT_BROWSERS_PATH"] = str(bundled)


def _console_utf8() -> None:
    if sys.platform == "win32":
        try:
            sys.stdout.reconfigure(encoding="utf-8")
            sys.stderr.reconfigure(encoding="utf-8")
        except Exception:
            pass


def _launch_browser(playwright):
    """Как в avito-parser-exactly.py: системный Chrome, при отсутствии — bundled Chromium."""
    try:
        return playwright.chromium.launch(channel="chrome", headless=False, args=CHROME_ARGS)
    except Exception:
        return playwright.chromium.launch(headless=False, args=CHROME_ARGS)


def _new_page(browser):
    """Контекст + stealth как в avito-parser-exactly.py (viewport 1920×1080, ru, Москва)."""
    context = browser.new_context(
        viewport={"width": 1920, "height": 1080},
        locale="ru-RU",
        timezone_id="Europe/Moscow",
    )
    stealth = Stealth(
        navigator_languages_override=("ru-RU", "ru", "en-US", "en"),
        navigator_platform_override="Win32",
        webgl_vendor_override="Intel Inc.",
        webgl_renderer_override="Intel Iris OpenGL Engine",
    )
    stealth.apply_stealth_sync(context)
    page = context.new_page()
    page.set_default_navigation_timeout(NAV_TIMEOUT_MS)
    page.set_default_timeout(25000)
    return context, page


def _human_pause(page, min_seconds: float = 0.4, max_seconds: float = 1.6) -> None:
    wait_ms = int(random.uniform(min_seconds, max_seconds) * 1000)
    page.wait_for_timeout(wait_ms)


def _simulate_human_activity(page) -> None:
    # Лёгкие случайные движения мышью — убираем "идеально ровный" машинный паттерн.
    try:
        for _ in range(random.randint(2, 4)):
            x = random.randint(100, 1200)
            y = random.randint(120, 760)
            page.mouse.move(x, y, steps=random.randint(8, 20))
            _human_pause(page, 0.1, 0.35)
    except Exception:
        pass


def _rewrite_css_urls(css: str, stylesheet_url: str) -> str:
    """Делает относительные url(...) в CSS абсолютными (нужно для инлайна стилей)."""
    base_p = urlparse(stylesheet_url)

    def repl(match: re.Match[str]) -> str:
        full = match.group(0)
        inner_raw = match.group(1).strip()
        inner = inner_raw.strip()
        quote = ""
        if len(inner) >= 2 and inner[0] in "'\"" and inner[-1] == inner[0]:
            quote = inner[0]
            inner = inner[1:-1].strip()
        u = inner
        if not u or u.startswith("data:") or u.startswith("#"):
            return full
        if u.startswith(("http://", "https://")):
            return full
        if u.startswith("//"):
            scheme = base_p.scheme or "https"
            resolved = f"{scheme}:{u}"
        elif u.startswith("/"):
            resolved = f"{base_p.scheme}://{base_p.netloc}{u}"
        else:
            resolved = urljoin(stylesheet_url, u)
        if quote:
            return f"url({quote}{resolved}{quote})"
        return f"url({resolved})"

    return re.sub(r"url\(\s*([^)]+)\s*\)", repl, css, flags=re.IGNORECASE)


def _inline_external_stylesheets(page) -> None:
    """
    Добавляет в <head> копию всех внешних CSS как один <style> (для офлайн-просмотра дампа).

    Важно: оригинальные <link rel="stylesheet"> не трогаем — иначе в окне Playwright
    страница теряет стили до перезагрузки (и порядок подключения ломается).
    """
    hrefs: list[str] = page.evaluate(
        """() => {
            const out = [];
            for (const link of document.querySelectorAll('link[rel="stylesheet"][href]')) {
                out.push(link.href);
            }
            return out;
        }"""
    )
    if not hrefs:
        return

    ordered_unique: list[str] = []
    seen: set[str] = set()
    for h in hrefs:
        if h not in seen:
            seen.add(h)
            ordered_unique.append(h)

    referer = str(page.url) if page.url else "https://www.avito.ru/"
    req_headers = {
        "Referer": referer,
        "Accept": "text/css,*/*;q=0.1",
    }

    chunks: list[str] = []
    for href in ordered_unique:
        try:
            response = page.request.get(href, timeout=25000, headers=req_headers)
            if not response.ok:
                continue
            css = response.text()
            css = _rewrite_css_urls(css, href)
            chunks.append(f"/* source: {href} */\n{css}\n")
        except Exception:
            continue

    if not chunks:
        return

    bundle = "\n".join(chunks)
    page.evaluate(
        """(cssText) => {
            for (const old of document.querySelectorAll("style[data-avito-dump-inlined]")) {
                old.remove();
            }
            const tag = document.createElement("style");
            tag.setAttribute("data-avito-dump-inlined", "1");
            tag.textContent = cssText;
            document.head.appendChild(tag);
        }""",
        bundle,
    )


def _wait_for_item_spa_ready(page, idx: int, total: int) -> None:
    """Ждём, пока карточка объявления отрисуется (а не дефолтная оболочка Авито)."""
    selectors = (
        '[data-marker="item-view/item-price"], '
        'h1[itemprop="name"], '
        '[data-marker="item-view/title-info"] h1'
    )
    try:
        page.locator(selectors).first.wait_for(state="visible", timeout=65000)
    except Exception as exc:
        print(f"[{idx}/{total}] Предупреждение: долго не появлялась карточка объявления: {exc}")


def _scroll_through_page(page, total_scrolls: int = 6) -> None:
    """Прокручиваем страницу постепенно: вниз пачкой, иногда чуть обратно, в конце — наверх."""
    try:
        size = page.viewport_size or {"height": 900}
        viewport_h = size.get("height") or 900
        for _ in range(total_scrolls):
            step = int(viewport_h * random.uniform(0.45, 0.9))
            page.mouse.wheel(0, step)
            _human_pause(page, 0.5, 1.3)
            if random.random() < 0.3:
                page.mouse.wheel(0, -int(step * random.uniform(0.2, 0.45)))
                _human_pause(page, 0.3, 0.8)
        page.evaluate("window.scrollTo({ top: 0, behavior: 'smooth' })")
        _human_pause(page, 0.8, 1.6)
    except Exception:
        pass


def _safe_slug(value: str) -> str:
    cleaned = "".join(ch if ch.isalnum() else "_" for ch in value.strip().lower())
    cleaned = "_".join(part for part in cleaned.split("_") if part)
    return cleaned[:80] or "house"


def _normalize_space(value: str) -> str:
    return " ".join((value or "").split()).strip()


def _load_urls(base_dir: Path) -> list[str]:
    urls_path = base_dir / URLS_FILE_NAME
    if not urls_path.exists():
        raise FileNotFoundError(
            f"Не найден файл со ссылками: {urls_path}. Создайте {URLS_FILE_NAME} рядом со скриптом."
        )
    urls: list[str] = []
    seen: set[str] = set()
    with open(urls_path, "r", encoding="utf-8") as f:
        for raw in f:
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            if line in seen:
                continue
            seen.add(line)
            urls.append(line)
    if not urls:
        raise ValueError(f"{URLS_FILE_NAME} не содержит ни одной валидной ссылки")
    return urls


def _extract_id_from_url(url: str) -> int | None:
    cleaned = url.split("?", 1)[0].rstrip("/")
    match = re.search(r"_(\d{6,})$", cleaned)
    if match:
        return int(match.group(1))
    return None


def _load_results_from_excel(base_dir: Path) -> list[dict]:
    """Подхватываем уже сохранённые строки из Excel, чтобы новый прогон их не затирал."""
    out_path = base_dir / RESULTS_XLSX_FILE_NAME
    if not out_path.exists():
        return []
    try:
        wb = load_workbook(out_path, read_only=True, data_only=True)
    except Exception:
        return []
    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header_row = next(it, None)
        if not header_row:
            return []
        header = [str(c).strip() if c is not None else "" for c in header_row]
        col_index = {name: i for i, name in enumerate(header)}
        out: list[dict] = []
        for row in it:
            if not row:
                continue
            rec: dict[str, str] = {}
            for col in EXPORT_COLUMNS:
                i = col_index.get(col)
                if i is not None and i < len(row):
                    v = row[i]
                    rec[col] = "" if v is None else str(v).strip()
                else:
                    rec[col] = ""
            if _normalize_space(rec.get("ссылка", "")):
                out.append(rec)
        return out
    except Exception:
        return []
    finally:
        wb.close()


def _write_excel(base_dir: Path, payload: list[dict]) -> None:
    out_path = base_dir / RESULTS_XLSX_FILE_NAME
    wb = Workbook()
    ws = wb.active
    ws.title = "Дома"
    ws.append(EXPORT_COLUMNS)
    for row in payload:
        ws.append([row.get(col, "") for col in EXPORT_COLUMNS])
    wb.save(out_path)


def _extract_title(page) -> str:
    for selector in (
        'h1[itemprop="name"]',
        '[data-marker="item-view/title-info"] h1',
        "h1",
    ):
        loc = page.locator(selector).first
        if loc.count() > 0:
            text = _normalize_space(loc.inner_text())
            if text:
                return text
    return ""


def _extract_price(page) -> str:
    candidates = (
        '[data-marker="item-view/item-price"]',
        'span[data-marker="item-view/item-price-content"]',
        '[itemprop="price"]',
    )
    for selector in candidates:
        loc = page.locator(selector).first
        if loc.count() == 0:
            continue
        try:
            content_attr = loc.get_attribute("content")
            if content_attr and re.search(r"\d", content_attr):
                return _normalize_space(content_attr)
        except Exception:
            pass
        text = _normalize_space(loc.inner_text())
        if text:
            return text
    return ""


def _extract_address(page) -> str:
    for selector in (
        '[itemprop="address"] span._8360df6eedcf8d52',
        '[itemprop="address"]',
        'div[class*="style-item-address"]',
    ):
        loc = page.locator(selector).first
        if loc.count() > 0:
            text = _normalize_space(loc.inner_text())
            if text:
                return text
    return ""


def _extract_description(page) -> str:
    loc = page.locator('div[data-marker="item-view/item-description"]').first
    if loc.count() == 0:
        return ""
    return _normalize_space(loc.inner_text())


def _extract_details_map(page) -> dict[str, str]:
    """Достаём пары `параметр: значение` из блока характеристик."""
    details: dict[str, str] = {}

    rows = page.locator("li.d2936d013c910379")
    if rows.count() == 0:
        # Резервный селектор — на случай, если CSS-хеши Авито поменялись.
        rows = page.locator(
            '[data-marker="item-view/params"] li, '
            'ul[class*="params"] li, '
            'div[class*="params"] li'
        )

    count = rows.count()
    for i in range(count):
        row = rows.nth(i)
        label_locator = row.locator(
            'span.d6e8fd2e3d52b32a, span[class*="styles-module-label"], span[class*="param-label"]'
        ).first
        full_text = _normalize_space(row.inner_text())
        label = ""
        value = ""
        if label_locator.count() > 0:
            label_text = _normalize_space(label_locator.inner_text())
            label = label_text.rstrip(":").lower()
            value = full_text.replace(label_text, "", 1).strip(" :")
        elif ":" in full_text:
            head, tail = full_text.split(":", 1)
            label = _normalize_space(head).lower()
            value = _normalize_space(tail)
        if label:
            details[label] = _normalize_space(value)
    return details


def _pick_house_field(details: dict[str, str], target_key: str) -> str:
    for label, value in details.items():
        for kw in HOUSE_FIELD_KEYWORDS.get(target_key, [target_key]):
            if kw in label:
                return value
    return ""


def _save_phone_image_from_popup(
    page,
    save_dir: Path,
    item_id: int,
    title: str,
    timestamp: str,
) -> str | None:
    button = page.locator('button[data-marker="item-phone-button/card"]').first
    if button.count() == 0:
        return None
    button.click(timeout=10000)

    popup = page.locator("div.ea88242d5926e753.b0253be71a159c76").first
    popup.wait_for(state="visible", timeout=10000)

    image = popup.locator("img._71e723cdca9c6624").first
    image.wait_for(state="visible", timeout=10000)
    src = image.get_attribute("src")
    if not src:
        return None

    content: bytes
    extension = "png"

    if src.startswith("data:image/"):
        header, _, encoded = src.partition(",")
        match = re.match(r"^data:image/([a-zA-Z0-9.+-]+);base64$", header)
        if not match:
            return None
        extension = match.group(1).replace("+xml", "")
        content = base64.b64decode(encoded)
    else:
        if src.startswith("//"):
            src = f"https:{src}"
        response = page.request.get(src, timeout=15000)
        if not response.ok:
            return None
        content_type = (response.header_value("content-type") or "").lower()
        match = re.search(r"image/([a-z0-9.+-]+)", content_type)
        if match:
            extension = match.group(1).replace("+xml", "")
        content = response.body()

    file_name = f"{item_id}_{_safe_slug(title)}_{timestamp}_phone.{extension}"
    out_path = save_dir / file_name
    with open(out_path, "wb") as f:
        f.write(content)
    return str(out_path)


def _extract_russian_phone(text: str) -> str | None:
    prepared = (
        (text or "")
        .replace("O", "0")
        .replace("o", "0")
        .replace("I", "1")
        .replace("l", "1")
        .replace("S", "5")
        .replace("B", "8")
    )
    cleaned = re.sub(r"[^\d+]", "", prepared)
    candidates: list[str] = []
    if cleaned:
        candidates.append(cleaned)
    candidates.extend(re.findall(r"(?:\+?7|8)\D*\d\D*\d\D*\d\D*\d\D*\d\D*\d\D*\d\D*\d\D*\d\D*\d", text))

    for candidate in candidates:
        digits = re.sub(r"\D", "", candidate)
        if len(digits) == 11 and (digits.startswith("8") or digits.startswith("7")):
            if digits.startswith("7"):
                digits = "8" + digits[1:]
            return digits
    return None


def _ocr_phone_number(image_path: str) -> str | None:
    with Image.open(image_path) as source_image:
        rgba = source_image.convert("RGBA")
        white_bg = Image.new("RGBA", rgba.size, "WHITE")
        white_bg.alpha_composite(rgba)
        image = ImageOps.grayscale(white_bg)
        # Увеличиваем и бинаризуем: для крупных чёрных цифр на белом фоне сильно повышает OCR-точность.
        image = image.resize((image.width * 3, image.height * 3))
        image = image.point(lambda p: 255 if p > 180 else 0)

    buf = io.BytesIO()
    image.save(buf, format="PNG")
    ocr_text = _OCR.classification(buf.getvalue())
    return _extract_russian_phone(ocr_text)


def _or_not_found(value: str | None) -> str:
    normalized = _normalize_space(value or "")
    return normalized if normalized else NOT_FOUND_VALUE


def _process_one_url(
    page,
    output_dir: Path,
    base_dir: Path,
    url: str,
    idx: int,
    total: int,
) -> dict | None:
    item_id = _extract_id_from_url(url) or idx
    print(f"[{idx}/{total}] Открываю: {url}")
    page.goto(url, wait_until="domcontentloaded", timeout=NAV_TIMEOUT_MS)
    _human_pause(page, 1.5, 4.0)
    _simulate_human_activity(page)
    _wait_for_item_spa_ready(page, idx, total)
    print(f"[{idx}/{total}] Жду {POST_GOTO_WAIT_S} сек. (как в avito-parser-exactly.py), затем прокручиваю страницу...")
    page.wait_for_timeout(POST_GOTO_WAIT_S * 1000)
    _scroll_through_page(page, total_scrolls=random.randint(5, 8))
    _simulate_human_activity(page)
    _human_pause(page, 1.0, 2.5)
    try:
        page.wait_for_load_state("networkidle", timeout=20000)
    except Exception:
        pass

    title = _extract_title(page) or f"house_{item_id}"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"{item_id}_{_safe_slug(title)}_{timestamp}.html"
    html_path = output_dir / file_name

    try:
        phone_image_path = _save_phone_image_from_popup(
            page=page,
            save_dir=base_dir,
            item_id=item_id,
            title=title,
            timestamp=timestamp,
        )
    except Exception as exc:
        print(f"[{idx}/{total}] id={item_id} Не удалось сохранить картинку телефона: {exc}")
        phone_image_path = None

    phone_number: str | None = None
    if phone_image_path:
        try:
            phone_number = _ocr_phone_number(phone_image_path)
        except Exception as exc:
            print(f"[{idx}/{total}] id={item_id} Ошибка OCR телефона: {exc}")
            phone_number = None
    if phone_image_path and phone_number:
        try:
            Path(phone_image_path).unlink(missing_ok=True)
            phone_image_path = None
        except Exception as exc:
            print(f"[{idx}/{total}] id={item_id} Не удалось удалить картинку телефона: {exc}")

    price = _extract_price(page)
    address = _extract_address(page)
    description = _extract_description(page)
    details = _extract_details_map(page)

    record = {
        "номер": item_id,
        "название": _or_not_found(title),
        "цена": _or_not_found(price),
        "адрес": _or_not_found(address),
        "телефон": phone_number if phone_number else (PHONE_NOT_RECOGNIZED if phone_image_path else NOT_FOUND_VALUE),
        "площадь дома": _or_not_found(_pick_house_field(details, "площадь дома")),
        "площадь участка": _or_not_found(_pick_house_field(details, "площадь участка")),
        "этажей": _or_not_found(_pick_house_field(details, "этажей")),
        "материал стен": _or_not_found(_pick_house_field(details, "материал стен")),
        "год постройки": _or_not_found(_pick_house_field(details, "год постройки")),
        "описание": _or_not_found(description),
        "характеристики": _or_not_found(json.dumps(details, ensure_ascii=False)) if details else NOT_FOUND_VALUE,
        "ссылка": url,
    }

    try:
        _inline_external_stylesheets(page)
        html_content = page.content()
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html_content)
    except Exception as exc:
        print(f"[{idx}/{total}] id={item_id} Не удалось сохранить HTML: {exc}")

    print(f"[{idx}/{total}] Сохранено: {html_path.name}")
    return record


def main() -> None:
    _setup_playwright_env()

    base_dir = project_root()
    output_dir = base_dir / OUTPUT_DIR_NAME
    output_dir.mkdir(parents=True, exist_ok=True)

    urls = _load_urls(base_dir)
    print(f"К обработке {len(urls)} ссылок.")

    results: list[dict] = _load_results_from_excel(base_dir)

    with sync_playwright() as p:
        browser = _launch_browser(p)
        context, page = _new_page(browser)
        try:
            total = len(urls)
            first_visit = True
            for idx, url in enumerate(urls, start=1):
                if not first_visit:
                    pause_s = BETWEEN_SITES_PAUSE_S + random.randint(0, 10)
                    print(f"[{idx}/{total}] Пауза между сайтами: {pause_s} сек.")
                    page.wait_for_timeout(pause_s * 1000)
                first_visit = False

                try:
                    record = _process_one_url(page, output_dir, base_dir, url, idx, total)
                except Exception as exc:
                    print(f"[{idx}/{total}] Ошибка обработки {url}: {exc}")
                    traceback.print_exc()
                    record = None

                if record is None:
                    continue

                replaced = False
                for i, existing in enumerate(results):
                    if (
                        isinstance(existing, dict)
                        and _normalize_space(str(existing.get("ссылка", ""))) == url
                    ):
                        results[i] = record
                        replaced = True
                        break
                if not replaced:
                    results.append(record)

                _write_excel(base_dir, results)

            print(f"Excel сохранён: {base_dir / RESULTS_XLSX_FILE_NAME}")
        finally:
            try:
                context.close()
            except Exception:
                pass
            try:
                browser.close()
            except Exception:
                pass


if __name__ == "__main__":
    _console_utf8()
    try:
        main()
    except Exception:
        print("Произошла ошибка:")
        traceback.print_exc()
        sys.exit(1)

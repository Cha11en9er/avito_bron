from __future__ import annotations

import base64
import io
import json
import os
import random
import re
import sys
import traceback
from datetime import datetime
from pathlib import Path

import ddddocr
from PIL import Image, ImageOps
from openpyxl import Workbook, load_workbook
from playwright.sync_api import sync_playwright
from playwright_stealth.stealth import Stealth

NAV_TIMEOUT_MS = 90000
WAIT_SECONDS = 35
MAX_COMPANIES = 1800
START_FROM_CARD_ID = 379
CARDS_FILE_NAME = "cards.json"
OUTPUT_DIR_NAME = "avito_exactly_dump"
RESULTS_FILE_NAME = "avito_exactly_results.json"
RESULTS_XLSX_FILE_NAME = "компании распил лдсп.xlsx"
NOT_FOUND_VALUE = "нету на сайте"
PHONE_NOT_RECOGNIZED = "не распознан"
_OCR = ddddocr.DdddOcr(show_ad=False)
EXPORT_COLUMNS = [
    "номер",
    "телефон",
    "адрес",
    "доставка",
    "оплата",
    "рейтинг",
    "количество отзывов",
    "описание",
    "название",
    "ссылка",
]


def _setup_playwright_env() -> None:
    if getattr(sys, "frozen", False):
        base = Path(sys.executable).resolve().parent
    else:
        base = Path(__file__).resolve().parent
    bundled = base / "ms-playwright"
    if bundled.is_dir():
        os.environ["PLAYWRIGHT_BROWSERS_PATH"] = str(bundled)


def _console_utf8() -> None:
    if sys.platform == "win32":
        try:
            sys.stdout.reconfigure(encoding="utf-8")
            sys.stderr.reconfigure(encoding="utf-8")
        except Exception:
            pass


def _launch_browser(playwright):
    try:
        return playwright.chromium.launch(channel="chrome", headless=False, args=CHROME_ARGS)
    except Exception:
        return playwright.chromium.launch(headless=False, args=CHROME_ARGS)


CHROME_ARGS = [
    "--start-maximized",
    "--disable-blink-features=AutomationControlled",
    "--disable-dev-shm-usage",
    "--no-first-run",
    "--disable-features=IsolateOrigins,site-per-process",
]

HUMAN_USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/133.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
]


def _new_page(browser):
    user_agent = random.choice(HUMAN_USER_AGENTS)
    viewport_w = random.choice([1920, 1912, 1904])
    viewport_h = random.choice([1080, 1032, 1040, 1008])
    context = browser.new_context(
        viewport={"width": viewport_w, "height": viewport_h},
        user_agent=user_agent,
        locale="ru-RU",
        timezone_id="Europe/Moscow",
        color_scheme="light",
        device_scale_factor=1,
    )
    try:
        stealth = Stealth(
            navigator_languages_override=("ru-RU", "ru", "en-US", "en"),
            navigator_platform_override="Win32",
            webgl_vendor_override="Intel Inc.",
            webgl_renderer_override="Intel Iris OpenGL Engine",
            navigator_user_agent_override=user_agent,
        )
    except TypeError:
        # Совместимость со старыми версиями playwright-stealth.
        stealth = Stealth(
            navigator_languages_override=("ru-RU", "ru", "en-US", "en"),
            navigator_platform_override="Win32",
            webgl_vendor_override="Intel Inc.",
            webgl_renderer_override="Intel Iris OpenGL Engine",
        )
    stealth.apply_stealth_sync(context)
    context.set_extra_http_headers(
        {
            "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
            "Upgrade-Insecure-Requests": "1",
            "Sec-CH-UA-Platform": '"Windows"',
            "Sec-CH-UA-Mobile": "?0",
        }
    )
    page = context.new_page()
    page.set_default_navigation_timeout(NAV_TIMEOUT_MS)
    page.set_default_timeout(25000)
    return context, page


def _human_pause(page, min_seconds: float = 0.4, max_seconds: float = 1.6) -> None:
    wait_ms = int(random.uniform(min_seconds, max_seconds) * 1000)
    page.wait_for_timeout(wait_ms)


def _simulate_human_activity(page) -> None:
    # Лёгкие действия, чтобы убрать "идеально ровный" машинный паттерн.
    try:
        for _ in range(random.randint(1, 2)):
            x = random.randint(100, 1200)
            y = random.randint(120, 760)
            page.mouse.move(x, y, steps=random.randint(8, 20))
            _human_pause(page, 0.1, 0.35)
        if random.random() < 0.85:
            scroll = random.randint(180, 900)
            page.mouse.wheel(0, scroll)
            _human_pause(page, 0.2, 0.7)
        if random.random() < 0.5:
            page.mouse.wheel(0, -random.randint(120, 500))
            _human_pause(page, 0.15, 0.55)
    except Exception:
        pass


def _safe_slug(value: str) -> str:
    cleaned = "".join(ch if ch.isalnum() else "_" for ch in value.strip().lower())
    cleaned = "_".join(part for part in cleaned.split("_") if part)
    return cleaned[:80] or "company"


def _load_cards(base_dir: Path) -> list[dict]:
    cards_path = base_dir / CARDS_FILE_NAME
    with open(cards_path, "r", encoding="utf-8") as f:
        cards = json.load(f)
    if not isinstance(cards, list):
        raise ValueError(f"{CARDS_FILE_NAME} должен содержать массив карточек")
    return cards


def _dedupe_cards_by_url(cards: list[dict]) -> list[dict]:
    seen_urls: set[str] = set()
    unique_cards: list[dict] = []
    for card in cards:
        url = _normalize_space(str(card.get("url", "")))
        if not url:
            continue
        if url in seen_urls:
            continue
        seen_urls.add(url)
        unique_cards.append(card)
    return unique_cards


def _write_results(base_dir: Path, payload: list[dict]) -> None:
    results_path = base_dir / RESULTS_FILE_NAME
    with open(results_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def _load_existing_results(base_dir: Path) -> list[dict]:
    results_path = base_dir / RESULTS_FILE_NAME
    if not results_path.exists():
        return []
    try:
        with open(results_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def _write_excel(base_dir: Path, payload: list[dict]) -> None:
    out_path = base_dir / RESULTS_XLSX_FILE_NAME
    wb = Workbook()
    ws = wb.active
    ws.title = "Компании"
    ws.append(EXPORT_COLUMNS)
    for row in payload:
        ws.append([row.get(col, "") for col in EXPORT_COLUMNS])
    wb.save(out_path)


def _load_existing_ids_from_excel(base_dir: Path) -> set[int]:
    out_path = base_dir / RESULTS_XLSX_FILE_NAME
    if not out_path.exists():
        return set()
    try:
        wb = load_workbook(out_path, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(min_row=2, values_only=True)
        ids: set[int] = set()
        for row in rows:
            if not row:
                continue
            row_dict = {EXPORT_COLUMNS[i]: (row[i] if i < len(row) else "") for i in range(len(EXPORT_COLUMNS))}
            if not _is_record_complete(row_dict):
                continue
            if not _is_record_quality_good(row_dict):
                continue
            value = row_dict.get("номер")
            if isinstance(value, int):
                ids.add(value)
            elif isinstance(value, str) and value.strip().isdigit():
                ids.add(int(value.strip()))
        wb.close()
        return ids
    except Exception:
        return set()


def _load_existing_links_from_excel(base_dir: Path) -> set[str]:
    out_path = base_dir / RESULTS_XLSX_FILE_NAME
    if not out_path.exists():
        return set()
    try:
        wb = load_workbook(out_path, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(min_row=2, values_only=True)
        links: set[str] = set()
        link_idx = EXPORT_COLUMNS.index("ссылка")
        for row in rows:
            if not row or link_idx >= len(row):
                continue
            row_dict = {EXPORT_COLUMNS[i]: (row[i] if i < len(row) else "") for i in range(len(EXPORT_COLUMNS))}
            if not _is_record_complete(row_dict):
                continue
            if not _is_record_quality_good(row_dict):
                continue
            value = _normalize_space(str(row[link_idx] or ""))
            if value:
                links.add(value)
        wb.close()
        return links
    except Exception:
        return set()


def _is_record_complete(record: dict) -> bool:
    for key in EXPORT_COLUMNS:
        if key not in record:
            return False
        value = _normalize_space(str(record.get(key, "")))
        if not value:
            return False
    return True


def _is_record_quality_good(record: dict) -> bool:
    """
    "Хорошая" запись = хотя бы одно из ключевых полей реально извлечено.
    Заглушки типа `нету на сайте` / `не распознан` считаем признаком капчи/ошибки.
    """
    link = _normalize_space(str(record.get("ссылка", "")))
    if not link:
        return False

    placeholder_values = {NOT_FOUND_VALUE, PHONE_NOT_RECOGNIZED}
    key_fields = ["адрес", "доставка", "оплата", "рейтинг", "количество отзывов", "описание", "телефон"]
    for field in key_fields:
        val = _normalize_space(str(record.get(field, "")))
        if val and val not in placeholder_values:
            return True
    return False


def _record_number(value) -> int | None:
    if isinstance(value, int):
        return value
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    return None


def _dedupe_results_by_link(payload: list[dict]) -> list[dict]:
    unique: list[dict] = []
    seen_links: set[str] = set()
    for item in sorted(payload, key=lambda x: _record_number(x.get("номер")) if isinstance(x, dict) else 10**9):
        if not isinstance(item, dict):
            continue
        link = _normalize_space(str(item.get("ссылка", "")))
        if not link:
            continue
        if link in seen_links:
            continue
        seen_links.add(link)
        unique.append(item)
    return unique


def _normalize_space(value: str) -> str:
    return " ".join((value or "").split()).strip()


def _extract_address(page) -> str | None:
    address_locator = page.locator('[itemprop="address"] span._8360df6eedcf8d52').first
    if address_locator.count() == 0:
        return None
    return _normalize_space(address_locator.inner_text())


def _extract_details_map(page) -> dict[str, str]:
    details: dict[str, str] = {}
    rows = page.locator("li.d2936d013c910379")
    count = rows.count()
    for i in range(count):
        row = rows.nth(i)
        label_locator = row.locator("span.d6e8fd2e3d52b32a").first
        if label_locator.count() == 0:
            continue
        label = _normalize_space(label_locator.inner_text()).rstrip(":").lower()
        full_text = _normalize_space(row.inner_text())
        value = full_text.replace(_normalize_space(label_locator.inner_text()), "", 1).strip(" :")
        if label:
            details[label] = _normalize_space(value)
    return details


def _extract_delivery(page, details: dict[str, str], description: str) -> str | None:
    delivery_all_russia = details.get("доставка по всей россии", "").lower()
    if delivery_all_russia == "нет":
        return "нету"

    delivery_value = details.get("доставка", "").lower()
    if delivery_value:
        if "нет" in delivery_value:
            return "нету"
        return "есть"

    additional_value = details.get("дополнительно", "").lower()
    if "доставка" in additional_value:
        return "есть"

    if "доставка" in description.lower():
        return "есть"
    return None


def _extract_payment(details: dict[str, str]) -> str | None:
    value = details.get("оплата")
    return _normalize_space(value) if value else None


def _extract_rating_and_reviews(page) -> tuple[str | None, str | None]:
    text_nodes = page.locator("span._3beac4bf2065c032.b77ace83bc35f090")
    total = text_nodes.count()
    rating = None
    reviews = None
    for i in range(total):
        text = _normalize_space(text_nodes.nth(i).inner_text())
        if rating is None and re.fullmatch(r"\d+(?:[,.]\d+)?", text):
            rating = text.replace(",", ".")
            continue
        if reviews is None:
            match = re.search(r"(\d+)\s+отзыв", text.lower())
            if match:
                reviews = match.group(1)
        if rating is not None and reviews is not None:
            break
    return rating, reviews


def _extract_description(page) -> str:
    description_locator = page.locator('div[data-marker="item-view/item-description"]').first
    if description_locator.count() == 0:
        return ""
    return description_locator.inner_text().strip()


def _save_phone_image_from_popup(
    page,
    save_dir: Path,
    item_id: int,
    page_number,
    title: str,
    timestamp: str,
) -> str | None:
    button = page.locator('button[data-marker="item-phone-button/card"]').first
    if button.count() == 0:
        return None
    button.click(timeout=10000)

    popup = page.locator("div.ea88242d5926e753.b0253be71a159c76").first
    popup.wait_for(state="visible", timeout=10000)

    image = popup.locator("img._71e723cdca9c6624").first
    image.wait_for(state="visible", timeout=10000)
    src = image.get_attribute("src")
    if not src:
        return None

    content: bytes
    extension = "png"

    if src.startswith("data:image/"):
        header, _, encoded = src.partition(",")
        match = re.match(r"^data:image/([a-zA-Z0-9.+-]+);base64$", header)
        if not match:
            return None
        extension = match.group(1).replace("+xml", "")
        content = base64.b64decode(encoded)
    else:
        if src.startswith("//"):
            src = f"https:{src}"
        response = page.request.get(src, timeout=15000)
        if not response.ok:
            return None
        content_type = (response.header_value("content-type") or "").lower()
        match = re.search(r"image/([a-z0-9.+-]+)", content_type)
        if match:
            extension = match.group(1).replace("+xml", "")
        content = response.body()

    page_part = f"page_{page_number}" if page_number is not None else "page_unknown"
    file_name = f"{item_id:02d}_{page_part}_{_safe_slug(title)}_{timestamp}_phone.{extension}"
    out_path = save_dir / file_name
    with open(out_path, "wb") as f:
        f.write(content)
    return str(out_path)


def _extract_russian_phone(text: str) -> str | None:
    prepared = (
        (text or "")
        .replace("O", "0")
        .replace("o", "0")
        .replace("I", "1")
        .replace("l", "1")
        .replace("S", "5")
        .replace("B", "8")
    )
    cleaned = re.sub(r"[^\d+]", "", prepared)
    candidates: list[str] = []
    if cleaned:
        candidates.append(cleaned)
    candidates.extend(re.findall(r"(?:\+?7|8)\D*\d\D*\d\D*\d\D*\d\D*\d\D*\d\D*\d\D*\d\D*\d\D*\d", text))

    for candidate in candidates:
        digits = re.sub(r"\D", "", candidate)
        if len(digits) == 11 and (digits.startswith("8") or digits.startswith("7")):
            if digits.startswith("7"):
                digits = "8" + digits[1:]
            return digits
    return None


def _ocr_phone_number(image_path: str) -> str | None:
    with Image.open(image_path) as source_image:
        rgba = source_image.convert("RGBA")
        white_bg = Image.new("RGBA", rgba.size, "WHITE")
        white_bg.alpha_composite(rgba)
        image = ImageOps.grayscale(white_bg)
        # Увеличиваем и бинаризуем: для крупных чёрных цифр на белом фоне это сильно повышает OCR-точность.
        image = image.resize((image.width * 3, image.height * 3))
        image = image.point(lambda p: 255 if p > 180 else 0)

    buf = io.BytesIO()
    image.save(buf, format="PNG")
    ocr_text = _OCR.classification(buf.getvalue())
    return _extract_russian_phone(ocr_text)


def _or_not_found(value: str | None) -> str:
    normalized = _normalize_space(value or "")
    return normalized if normalized else NOT_FOUND_VALUE


def main() -> None:
    _setup_playwright_env()
    from playwright.sync_api import sync_playwright

    base_dir = Path(__file__).resolve().parent
    output_dir = base_dir / OUTPUT_DIR_NAME
    output_dir.mkdir(parents=True, exist_ok=True)
    cards = _dedupe_cards_by_url(_load_cards(base_dir))
    cards = [card for card in cards if (_record_number(card.get("id")) or 0) >= START_FROM_CARD_ID]
    cards = cards[:MAX_COMPANIES]
    if not cards:
        raise ValueError(f"В cards.json нет карточек для обработки c id >= {START_FROM_CARD_ID}")
    print(f"Старт обработки карточек с id >= {START_FROM_CARD_ID}. Карточек к обходу: {len(cards)}")
    results: list[dict] = _load_existing_results(base_dir)
    parsed_ids_json = {
        _record_number(item.get("номер"))
        for item in results
        if isinstance(item, dict) and _is_record_quality_good(item) and _record_number(item.get("номер")) is not None
    }
    parsed_links_json = {
        _normalize_space(str(item.get("ссылка", "")))
        for item in results
        if isinstance(item, dict) and _is_record_quality_good(item) and _normalize_space(str(item.get("ссылка", "")))
    }
    parsed_ids_excel = _load_existing_ids_from_excel(base_dir)
    parsed_links_excel = _load_existing_links_from_excel(base_dir)
    parsed_ids = parsed_ids_json | parsed_ids_excel
    parsed_links = parsed_links_json | parsed_links_excel

    with sync_playwright() as p:
        browser = _launch_browser(p)
        context, page = _new_page(browser)
        try:
            total = len(cards)
            for idx, card in enumerate(cards, start=1):
                card_id = _record_number(card.get("id"))
                if card_id is None:
                    print(f"[{idx}/{total}] Пропуск: в cards.json нет валидного id.")
                    continue

                title = str(card.get("title") or f"company_{card_id}")
                url = str(card.get("url") or "").strip()
                page_number = card.get("page")
                if not url:
                    print(f"[{idx}/{total}] Пропуск: пустой url у карточки '{title}'")
                    continue
                if card_id in parsed_ids or url in parsed_links:
                    print(f"[{idx}/{total}] id={card_id} ссылка уже обработана, пропускаю без открытия.")
                    continue

                print(f"[{idx}/{total}] Открываю: {title}")
                page.goto(url, wait_until="domcontentloaded", timeout=NAV_TIMEOUT_MS)
                _human_pause(page, 1.5, 4.0)
                _simulate_human_activity(page)
                dynamic_wait = WAIT_SECONDS + random.randint(3, 14)
                print(f"[{idx}/{total}] Жду {dynamic_wait} секунд...")
                page.wait_for_timeout(dynamic_wait * 1000)

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                page_part = f"page_{page_number}" if page_number is not None else "page_unknown"
                file_name = f"{card_id:02d}_{page_part}_{_safe_slug(title)}_{timestamp}.html"
                out_path = output_dir / file_name

                html_content = page.content()
                with open(out_path, "w", encoding="utf-8") as f:
                    f.write(html_content)

                try:
                    phone_image_path = _save_phone_image_from_popup(
                        page=page,
                        save_dir=Path.cwd(),
                        item_id=card_id,
                        page_number=page_number,
                        title=title,
                        timestamp=timestamp,
                    )
                except Exception as exc:
                    print(f"[{idx}/{total}] id={card_id} Не удалось сохранить изображение телефона: {exc}")
                    phone_image_path = None
                phone_number = None
                if phone_image_path:
                    try:
                        phone_number = _ocr_phone_number(phone_image_path)
                    except Exception as exc:
                        print(f"[{idx}/{total}] id={card_id} Ошибка OCR телефона: {exc}")
                        phone_number = None
                if phone_image_path and phone_number:
                    try:
                        Path(phone_image_path).unlink(missing_ok=True)
                        phone_image_path = None
                    except Exception as exc:
                        print(f"[{idx}/{total}] id={card_id} Не удалось удалить картинку телефона: {exc}")

                description = _extract_description(page)
                details = _extract_details_map(page)
                address = _extract_address(page)
                delivery = _extract_delivery(page, details, description)
                payment = _extract_payment(details)
                rating, reviews_count = _extract_rating_and_reviews(page)

                result = {
                    "номер": card_id,
                    "название": title,
                    "ссылка": url,
                    "адрес": _or_not_found(address),
                    "доставка": _or_not_found(delivery),
                    "оплата": _or_not_found(payment),
                    "рейтинг": _or_not_found(rating),
                    "количество отзывов": _or_not_found(reviews_count),
                    "описание": _or_not_found(description),
                    "телефон": phone_number if phone_number else (PHONE_NOT_RECOGNIZED if phone_image_path else NOT_FOUND_VALUE),
                }

                replaced = False
                for i, existing in enumerate(results):
                    if isinstance(existing, dict) and _record_number(existing.get("номер")) == card_id:
                        results[i] = result
                        replaced = True
                        break
                if not replaced:
                    results.append(result)
                results = _dedupe_results_by_link(results)
                parsed_ids.add(card_id)
                parsed_links.add(url)
                _write_results(base_dir, results)
                _write_excel(base_dir, results)
                print(f"[{idx}/{total}] Сохранено: {out_path}")
            print(f"JSON сохранён: {base_dir / RESULTS_FILE_NAME}")
            print(f"Excel сохранён: {base_dir / RESULTS_XLSX_FILE_NAME}")
        finally:
            try:
                context.close()
            except Exception:
                pass
            browser.close()


if __name__ == "__main__":
    _console_utf8()
    try:
        main()
    except Exception:
        print("Произошла ошибка:")
        traceback.print_exc()
        sys.exit(1)